from openpyxl import load_workbook
p = r"c:/xampp/htdocs/UASPMS/database/imports/RPCPPE 2025.xlsx"
wb = load_workbook(p, data_only=True)
print("SHEETS:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    mr = ws.max_row
    mc = ws.max_column
    headers = [ws.cell(1, c).value for c in range(1, mc + 1)]
    print("---", s, "rows", mr, "cols", mc)
    print("HEADERS:", headers)
